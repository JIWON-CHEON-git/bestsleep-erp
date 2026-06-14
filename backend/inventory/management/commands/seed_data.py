"""초기 데이터 적재 명령어.

사용법:
    python manage.py seed_data --excel ../data/exampledata.xlsx
    python manage.py seed_data            # 엑셀 없이 하드코딩 fallback 데이터 사용

엑셀 파일이 없거나, 필요한 시트를 찾지 못하면 inventory/fallback_data.py에
정의된 하드코딩 데이터로 자동 대체한다.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from inventory import fallback_data
from inventory.models import (
    DailyLedger,
    Order,
    OrderFulfillment,
    Product,
    ProductionOrder,
    SimulationConfig,
)


def _parse_date(value: Any) -> Optional[date]:
    """엑셀 셀 값을 date로 변환. 비어 있으면 None."""

    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())


class Command(BaseCommand):
    help = "엑셀(또는 fallback 하드코딩 데이터)로부터 제품/주문/시뮬레이션 설정을 적재합니다."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "--excel",
            type=str,
            default=None,
            help="엑셀 파일 경로 (예: ../data/exampledata.xlsx). 미지정 시 fallback 데이터 사용.",
        )

    def handle(self, *args, **options) -> None:
        excel_path = options.get("excel")

        products: list[tuple]
        orders: list[tuple]
        sim_start: date
        sim_end: date

        if excel_path:
            path = Path(excel_path)
            if not path.exists():
                raise CommandError(f"엑셀 파일을 찾을 수 없습니다: {path}")
            products, orders, sim_start, sim_end = self._load_from_excel(path)
        else:
            self.stdout.write("엑셀 경로가 지정되지 않아 fallback 데이터를 사용합니다.")
            products, orders, sim_start, sim_end = self._load_fallback()

        with transaction.atomic():
            self._reset_tables()
            self._create_products(products)
            self._create_orders(orders)
            SimulationConfig.objects.create(start_date=sim_start, end_date=sim_end)

        self.stdout.write(
            self.style.SUCCESS(
                f"시드 완료: 제품 {len(products)}개, 주문 {len(orders)}건, "
                f"시뮬레이션 기간 {sim_start} ~ {sim_end}"
            )
        )

    # ------------------------------------------------------------------
    # 데이터 로드
    # ------------------------------------------------------------------

    def _load_fallback(self) -> tuple[list[tuple], list[tuple], date, date]:
        products = fallback_data.PRODUCTS
        orders = fallback_data.ORDERS
        sim_start = date.fromisoformat(fallback_data.SIMULATION_START_DATE)
        sim_end = date.fromisoformat(fallback_data.SIMULATION_END_DATE)
        return products, orders, sim_start, sim_end

    def _load_from_excel(self, path: Path) -> tuple[list[tuple], list[tuple], date, date]:
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - openpyxl is a hard dependency
            raise CommandError("openpyxl이 설치되어 있지 않습니다.") from exc

        wb = openpyxl.load_workbook(path, data_only=True)

        try:
            products = self._read_products_sheet(wb)
            orders = self._read_orders_sheet(wb)
            sim_start, sim_end = self._read_parameters_sheet(wb)
        except (KeyError, ValueError) as exc:
            self.stdout.write(
                self.style.WARNING(
                    f"엑셀 형식을 읽는 중 문제가 발생해 fallback 데이터를 사용합니다: {exc}"
                )
            )
            return self._load_fallback()

        return products, orders, sim_start, sim_end

    def _sheet_rows(self, wb, sheet_name: str) -> tuple[list[str], list[tuple]]:
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"시트를 찾을 수 없습니다: {sheet_name}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"시트가 비어 있습니다: {sheet_name}")
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = [r for r in rows[1:] if any(c is not None for c in r)]
        return header, data_rows

    def _read_products_sheet(self, wb) -> list[tuple]:
        header, rows = self._sheet_rows(wb, "제품마스터")
        idx = {name: i for i, name in enumerate(header)}

        required = [
            "SKU", "제품명", "카테고리", "사이즈", "단가(원)", "초기재고",
            "안전재고", "재주문점", "생산로트(MOQ)", "생산리드타임(일)", "배송리드타임(일)",
        ]
        for col in required:
            if col not in idx:
                raise KeyError(f"제품마스터 시트에 '{col}' 컬럼이 없습니다.")

        products = []
        for r in rows:
            products.append((
                str(r[idx["SKU"]]).strip(),
                str(r[idx["제품명"]]).strip(),
                str(r[idx["카테고리"]]).strip(),
                str(r[idx["사이즈"]]).strip(),
                int(r[idx["단가(원)"]]),
                int(r[idx["초기재고"]]),
                int(r[idx["안전재고"]]),
                int(r[idx["재주문점"]]),
                int(r[idx["생산로트(MOQ)"]]),
                int(r[idx["생산리드타임(일)"]]),
                int(r[idx["배송리드타임(일)"]]),
            ))
        return products

    def _read_orders_sheet(self, wb) -> list[tuple]:
        header, rows = self._sheet_rows(wb, "주문")
        idx = {name: i for i, name in enumerate(header)}

        required = ["주문번호", "주문일", "SKU", "수량", "고객명"]
        for col in required:
            if col not in idx:
                raise KeyError(f"주문 시트에 '{col}' 컬럼이 없습니다.")

        desired_col = "희망배송일(선택)" if "희망배송일(선택)" in idx else None

        orders = []
        for r in rows:
            desired = _parse_date(r[idx[desired_col]]) if desired_col else None
            orders.append((
                str(r[idx["주문번호"]]).strip(),
                _parse_date(r[idx["주문일"]]).isoformat(),
                str(r[idx["SKU"]]).strip(),
                int(r[idx["수량"]]),
                str(r[idx["고객명"]]).strip(),
                desired.isoformat() if desired else None,
            ))
        return orders

    def _read_parameters_sheet(self, wb) -> tuple[date, date]:
        header, rows = self._sheet_rows(wb, "파라미터")
        idx = {name: i for i, name in enumerate(header)}

        if "항목" not in idx or "값" not in idx:
            raise KeyError("파라미터 시트에 '항목'/'값' 컬럼이 없습니다.")

        params: dict[str, Any] = {}
        for r in rows:
            key = str(r[idx["항목"]]).strip()
            params[key] = r[idx["값"]]

        if "시뮬레이션_시작일" not in params or "시뮬레이션_종료일" not in params:
            raise KeyError("파라미터 시트에 시뮬레이션 시작일/종료일이 없습니다.")

        sim_start = _parse_date(params["시뮬레이션_시작일"])
        sim_end = _parse_date(params["시뮬레이션_종료일"])
        if sim_start is None or sim_end is None:
            raise ValueError("시뮬레이션 시작일/종료일을 해석할 수 없습니다.")
        return sim_start, sim_end

    # ------------------------------------------------------------------
    # DB 적재
    # ------------------------------------------------------------------

    def _reset_tables(self) -> None:
        OrderFulfillment.objects.all().delete()
        DailyLedger.objects.all().delete()
        ProductionOrder.objects.all().delete()
        Order.objects.all().delete()
        Product.objects.all().delete()
        SimulationConfig.objects.all().delete()

    def _create_products(self, products: list[tuple]) -> None:
        Product.objects.bulk_create([
            Product(
                sku=sku,
                name=name,
                category=category,
                size=size,
                unit_price=unit_price,
                initial_stock=initial_stock,
                safety_stock=safety_stock,
                reorder_point=reorder_point,
                moq=moq,
                production_lead_time_days=production_lt,
                delivery_lead_time_days=delivery_lt,
            )
            for (
                sku, name, category, size, unit_price, initial_stock,
                safety_stock, reorder_point, moq, production_lt, delivery_lt,
            ) in products
        ])

    def _create_orders(self, orders: list[tuple]) -> None:
        Order.objects.bulk_create([
            Order(
                order_no=order_no,
                order_date=date.fromisoformat(order_date),
                product_id=sku,
                quantity=quantity,
                customer_name=customer_name,
                desired_delivery_date=(
                    date.fromisoformat(desired) if desired else None
                ),
            )
            for (order_no, order_date, sku, quantity, customer_name, desired) in orders
        ])
